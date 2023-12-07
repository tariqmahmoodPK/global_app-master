import frappe
from openpyxl import load_workbook
from frappe.desk.form.linked_with import get_linked_docs, get_linked_doctypes

@frappe.whitelist()
def change_so_status(so):
    so_record = frappe.get_doc("Sales Order", so)
    so_linkinfo = get_linked_doctypes("Sales Order")
    so_docs = get_linked_docs("Sales Order", so, so_linkinfo)
    dn_docs = {}
    si_status = ""

    if "Delivery Note" in so_docs:
        dn_linkinfo = get_linked_doctypes("Delivery Note")
        dn_docs = get_linked_docs("Delivery Note", so_docs['Delivery Note'][0]['name'], dn_linkinfo)

    print(so_docs)
    #STOCK ORDER
    if "Sales Invoice" in so_docs:
        si_status = frappe.get_value("Sales Invoice", so_docs['Sales Invoice'][0]['name'], "status")

    if so_docs and "Delivery Note" in so_docs and "Purchase Order" not in so_docs and "Packing Slip" not in dn_docs:
        change(so,"DN Generated", "Stock Order")

    elif so_docs and "Delivery Note" in so_docs and "Purchase Order" not in so_docs and "Packing Slip" in dn_docs:
        change(so,"Shipping", "Stock Order")

    elif so_record.shipment_details and so_docs and "Delivery Note" in so_docs and "Purchase Order" not in so_docs and "Packing Slip" in dn_docs:
        change(so,"Shipped", "Stock Order")

    elif so_docs and so_record.shipment_details and "Sales Invoice" in so_docs and si_status == "Unpaid" and "Delivery Note" in so_docs and "Purchase Order" not in so_docs and "Packing Slip" in dn_docs:
        change(so,"Billed", "Stock Order")

    elif so_docs and so_record.shipment_details and "Sales Invoice" in so_docs and si_status == "Paid" and "Delivery Note" in so_docs and "Purchase Order" not in so_docs and "Packing Slip" in dn_docs:
        change(so,"Paid", "Stock Order")

    #To Vendor
    elif so_docs and "Delivery Note" not in so_docs and "Purchase Order" in so_docs and "Packing Slip" not in dn_docs:
        change(so,"PO Generated/ Vendor Emailed / Sent to Gas", "To Vendor")

    elif so_docs and so_record.shipment_details and "Delivery Note" not in so_docs and "Purchase Order" in so_docs and "Packing Slip" not in dn_docs:
        change(so,"Received TN and SC from Gas", "To Vendor")

    #Split Order - Stock Order
    elif so_docs and "Delivery Note" in so_docs and "Purchase Order" in so_docs and "Packing Slip" not in dn_docs:
        change(so,"PO/DN Generated", "Split Order")

    elif so_docs and "Delivery Note" in so_docs and "Purchase Order" in so_docs and "Packing Slip" in dn_docs:
        change(so,"Shipping", "Split Order")

    elif so_docs and so_record.shipment_details and "Delivery Note" in so_docs and "Purchase Order" in so_docs and "Packing Slip" in dn_docs:
        change(so,"Shipped", "Split Order")

    elif so_docs and so_record.shipment_details and "Sales Invoice" in so_docs and si_status == "Unpaid" and "Delivery Note" in so_docs and "Purchase Order" in so_docs and "Packing Slip" in dn_docs:
        change(so,"Billed", "Split Order")

    elif so_docs and so_record.shipment_details and "Sales Invoice" in so_docs and si_status == "Paid" and "Delivery Note" in so_docs and "Purchase Order" in so_docs and "Packing Slip" in dn_docs:
        change(so,"Paid", "Split Order")

    #Split - To Vendor
    elif so_docs and "Delivery Note" in so_docs and "Purchase Order" in so_docs and "Packing Slip" in dn_docs:
        change(so,"PO Generated/ Vendor Emailed / Sent to Gas", "Split Order")



def change(name, status, type_status):
    frappe.db.sql(""" UPDATE `tabSales Order` SET so_type=%s, so_type_status=%s WHERE name=%s """, (type_status, status,name))
    frappe.db.commit()

def load_excel():
    sales_order_with_lacking = frappe.db.sql(""" SELECT name, shipment_details FROM `tabSales Order` WHERE (shipment_cost is null OR shipment_cost = 0) and shipment_details != '' and transaction_date > %s and docstatus = 1""",'2020-07-19', as_dict=1)
    print(sales_order_with_lacking)
    #
    loc = ("gaserp.tailerp.com/public/files/EWS Export US Locale en Language.xlsx")
    wb2 = load_workbook(loc)
    ws = wb2.active
    counter = 2
    if len(sales_order_with_lacking) > 0:
        while ws['E' + str(counter)].value != None:
            ps = frappe.db.sql(""" SELECT parent FROM `tabPacking Slip packages shadow` WHERE shipment_tracking_number=%s """, ws['E' + str(counter)].value, as_dict=1)
            if len(ps) > 0:
                dn = frappe.db.sql(""" SELECT delivery_note FROM `tabPacking Slip` WHERE name=%s""",ps[0].parent,as_dict=1)
                if len(dn) > 0:
                    sales_order = frappe.db.sql("""
                                    SELECT SO.shipment_cost, SO.name FROM `tabDelivery Note Item` AS DNI
                                    INNER JOIN `tabSales Order` AS SO ON SO.name = DNI.against_sales_order and transaction_date > '2020-07-01'
                                    WHERE DNI.parent=%s LIMIT 1
                                    """, dn[0].delivery_note, as_dict=1)
                    print(sales_order)
                    if len(sales_order) > 0 and so_in_so_with_lacking(sales_order_with_lacking,sales_order[0].name):
                        print(sales_order[0].name)
                        if sales_order[0].shipment_cost and float(sales_order[0].shipment_cost) >= 0:
                            print("NAA DIRI SA FIRST")
                            frappe.db.sql(""" UPDATE `tabSales Order` SET shipment_cost=%s WHERE name=%s""", (str(float(sales_order[0].shipment_cost) + float(ws['J' + str(counter)].value)),sales_order[0].name))
                            frappe.db.commit()
                            print("Done " + sales_order[0].name)

                        elif not sales_order[0].shipment_cost:
                            print("NAA DIRI SA SECOND")

                            frappe.db.sql(""" UPDATE `tabSales Order` SET shipment_cost=%s  WHERE name=%s""", (str(float(ws['J' + str(counter)].value)),sales_order[0].name))
                            frappe.db.commit()
                            print("Done " + sales_order[0].name)

            # print(counter)
            # print(ws['E' + str(counter)].value)
            # print(ws['J' + str(counter)].value)
            counter += 1
def so_in_so_with_lacking(sales_orders, name):
    for i in sales_orders:
        if i.name == name:
            return True
    return False

def update_so_from_ps():
    sales_orders = frappe.db.sql(""" SELECT * FROM `tabSales Order` WHERE shipment_date is null and docstatus=1 and transaction_date > '2020-06-30'""",as_dict=1)
    for so in sales_orders:
        dn = frappe.db.sql(""" SELECT parent FROM `tabDelivery Note Item` WHERE against_sales_order=%s LIMIT 1""", so.name, as_dict=1)
        if len(dn) > 0:
            ps = frappe.db.sql(""" SELECT * FROM `tabPacking Slip` WHERE delivery_note=%s """, dn[0].parent,as_dict=1)
            if len(ps) > 0:
                packages = frappe.db.sql(""" SELECT * FROM `tabPacking Slip packages shadow` WHERE parent=%s""",ps[0].name,as_dict=1)
                tracking_number = ""
                for i in packages:
                    if i.shipment_tracking_number:
                        tracking_number += i.shipment_tracking_number

                frappe.db.sql(""" UPDATE `tabSales Order` SET shipment_details=%s, shipment_date=%s WHERE name=%s""", (tracking_number,ps[0].creation.date(), so.name))
                frappe.db.commit()
                print("DONE FOR " + so.name + ", " + ps[0].name)


def execute_creation_of_items():
    items = frappe.db.sql("""  SELECT * FROM `tabItem` WHERE sync_to_cobb=0""",as_dict=1)
    for i in items:
        print("ITEM")
        print(i.name)
        item = frappe.get_doc("Item", i.name)
        item.save()
        frappe.db.sql("""  UPDATE `tabItem` SET sync_to_cobb=1 WHERE name=%s""",(i.name), as_dict=1)
        frappe.db.commit()
        print("DOOOOOOOOONE")


def reset_stocks():
    for i in range(0,100):
        print(i)
        if len(get_items()) > 0:
            doc_se = {
                "doctype": "Stock Entry",
                "stock_entry_type": "Material Issue",
                "items": get_items()
            }
            print(doc_se)
            try:
                frappe.get_doc(doc_se).insert(ignore_permissions=1).submit()
            except:
                print(frappe.get_traceback())
        else:
            break
def get_items():
    from erpnext.stock.stock_ledger import get_previous_sle
    from erpnext.stock.utils import get_latest_stock_qty


    time = frappe.utils.now_datetime().time()
    date = frappe.utils.now_datetime().date()
    items = frappe.db.sql(""" SELECT * FROM `tabItem` where is_stock_item=1 and disabled=0""", as_dict=1)
    warehouses = frappe.db.sql(""" SELECT * FROM `tabWarehouse` where is_group=0""", as_dict=1)

    se_item = []
    count = 0
    for i in items:
        if count < 50:
            for ii in warehouses:
                previous_sle =  get_latest_stock_qty(i.name, ii.name)
                # print(previous_sle)
                if previous_sle and previous_sle > 0:
                    count += 1
                    # print( i.name + " " + str(previous_sle.get("qty_after_transaction")) + " " + previous_sle.get("warehouse"))
                    se_item.append({
                        'item_code': i.name,
                        's_warehouse': ii.name,
                        'qty': previous_sle,
                        'uom': i.stock_uom,
                        'cost_center': 'Main - Global',
                    })
        else:
            break
    return se_item

def get_single_item_items():
    from erpnext.stock.stock_ledger import get_previous_sle
    from erpnext.stock.utils import get_latest_stock_qty

    previous_sle = get_latest_stock_qty("NSFB-5058B", "Finished Goods - COBB")
    print(previous_sle)
    # time = frappe.utils.now_datetime().time()
    # date = frappe.utils.now_datetime().date()
    # items = frappe.db.sql(""" SELECT * FROM `tabItem`""", as_dict=1)
    # warehouses = frappe.db.sql(""" SELECT * FROM `tabWarehouse` where is_group=0""", as_dict=1)
    #
    # se_item = []
    # count = 0
    # for i in items:
    #     if count < 50:
    #         for ii in warehouses:
    #
    #             # print(previous_sle)
    #             if previous_sle and previous_sle < 0:
    #                 count += 1
    #                 # print( i.name + " " + str(previous_sle.get("qty_after_transaction")) + " " + previous_sle.get("warehouse"))
    #                 se_item.append({
    #                     'item_code': i.name,
    #                     't_warehouse': ii.name,
    #                     'qty': abs(previous_sle),
    #                     'uom': i.stock_uom,
    #                     'allow_zero_valuation_rate': 1,
    #                     'cost_center': 'Main - COBB',
    #                 })
    #     else:
    #         break
    # return se_item


def update_pick_list():
    packing_slip = frappe.db.sql(""" SELECT * FROM `tabPacking Slip` WHERE """)


def update_packing_slip_item_codes():
    ps = frappe.db.sql(""" SELECT * FROM `tabPacking Slip`  WHERE (item_codes='' or item_codes is null) or (packages_list='' or packages_list is null) or (total_item_qty = 0) """,as_dict=1)
    for i in ps:
        ps_doc = frappe.get_doc("Packing Slip", i.name)
        dn_doc = frappe.get_doc("Delivery Note", i.delivery_note)
        if not ps_doc.item_codes:
            codes = []
            for ii in dn_doc.items:
                codes.append(ii.item_code)
            codes.sort()
            frappe.db.sql(""" UPDATE `tabPacking Slip` SET item_codes=%s WHERE name=%s""",(",".join(codes), i.name))
            frappe.db.commit()
            print("DONE UPDATING ITEM CODES FOR "+i.name)
        if ps_doc.total_item_qty == 0:
            total_qty = 0
            for ii in dn_doc.items:
                total_qty += ii.qty

            frappe.db.sql(""" UPDATE `tabPacking Slip` SET total_item_qty=%s WHERE name=%s""",
                          (total_qty, i.name))
            frappe.db.commit()
            print("DONE UPDATING TOTAL QTY FOR "+i.name)

        if ps_doc.packages_list != '{}' :
            packages = []
            weights = {}
            for iii in ps_doc.packages_information:
                if iii.items and iii.shipment_tracking_number:
                    packages.append(iii.package)
                    weights[iii.package] = iii.weight
            print(weights)
            packages.sort()
            print(i.name)
            frappe.db.sql(""" UPDATE `tabPacking Slip` SET packages_list=%s, weights=%s WHERE name=%s""",
                          (",".join(packages),str(weights), i.name))
            frappe.db.commit()
            print("DONE UPDATING PACKAGES LIST FOR " + i.name)


def update_packing_slip_weights():
    query = """ SELECT * FROM `tabPacking Slip` WHERE weights LIKE "%: '%" """
    ps = frappe.db.sql(query, as_dict=1)

    for i in ps:
        print(i)
        weights = {}
        for weight in eval(i.weights):
            weights[weight] = float(eval(i.weights)[weight])
        update_query = """ UPDATE `tabPacking Slip` SET weights="{0}" WHERE name='{1}'""".format(weights, i.name)
        print(update_query)
        frappe.db.sql(update_query)
        frappe.db.commit()
